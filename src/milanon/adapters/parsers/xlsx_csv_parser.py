"""XLSX/CSV parser — extracts text and structured data from spreadsheets."""

from __future__ import annotations

import csv
import io
import logging
from pathlib import Path

import openpyxl

from milanon.domain.entities import DocumentFormat, ExtractedDocument

logger = logging.getLogger(__name__)

# PISA 410 exports have a report title in row 1; row 2 is the header.
_PISA_410_COL_COUNT = 45


def _is_pisa_410(rows: list[list[str]]) -> bool:
    """Return True if the sheet looks like a PISA 410 export.

    Heuristic: at least 2 rows, the first row has exactly 1 non-empty cell
    (the report title), and the second row has >= 3 non-empty cells.
    """
    if len(rows) < 2:
        return False
    non_empty_first = sum(1 for cell in rows[0] if str(cell).strip())
    non_empty_second = sum(1 for cell in rows[1] if str(cell).strip())
    return non_empty_first == 1 and non_empty_second >= 3


def _rows_to_text(rows: list[list[str]]) -> str:
    """Render a 2-D list of strings as pipe-separated text lines."""
    lines = []
    for row in rows:
        lines.append(" | ".join(str(cell) for cell in row))
    return "\n".join(lines)


class XlsxCsvParser:
    """Parses XLSX and CSV files into ExtractedDocument.

    XLSX strategy:
    - Open with openpyxl (read-only, data-only).
    - Each sheet becomes one table in structured_content["tables"].
    - PISA 410 sheets (detected by heuristic): row 1 (title) is skipped,
      row 2 becomes the header.

    CSV strategy:
    - Auto-detect delimiter: try semicolon first, then comma.
    - Entire file is one table.

    text_content is a flat pipe-separated representation of all tables,
    separated by a Markdown HR for multi-sheet XLSX.
    """

    def parse(self, path: Path) -> ExtractedDocument:
        """Parse a spreadsheet file and return its extracted content."""
        suffix = path.suffix.lower()
        if suffix in (".xlsx", ".xls"):
            return self._parse_xlsx(path)
        if suffix == ".csv":
            return self._parse_csv(path)
        raise ValueError(f"Unsupported extension: {suffix}")

    def supported_extensions(self) -> list[str]:
        """Return the file extensions this parser handles."""
        return [".xlsx", ".xls", ".csv"]

    # ------------------------------------------------------------------
    # XLSX
    # ------------------------------------------------------------------

    def _parse_xlsx(self, path: Path) -> ExtractedDocument:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        tables: list[list[list[str]]] = []
        sheet_names: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            raw_rows = self._read_worksheet(ws)
            if not raw_rows:
                continue
            data_rows = self._apply_pisa_skip(raw_rows)
            tables.append(data_rows)
            sheet_names.append(sheet_name)

        wb.close()

        text_parts = [
            f"{name}\n{_rows_to_text(t)}" for name, t in zip(sheet_names, tables, strict=False)
        ]
        text_content = "\n\n---\n\n".join(text_parts)

        metadata: dict[str, str] = {
            "sheet_count": str(len(tables)),
            "sheet_names": ", ".join(sheet_names),
        }

        return ExtractedDocument(
            source_path=str(path),
            format=DocumentFormat.XLSX,
            text_content=text_content,
            structured_content={"tables": tables} if tables else None,
            metadata=metadata,
            embedded_image_count=0,
        )

    def _read_worksheet(self, ws) -> list[list[str]]:
        """Read all non-empty rows from a worksheet as string lists."""
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            str_row = [str(cell) if cell is not None else "" for cell in row]
            if any(cell.strip() for cell in str_row):
                rows.append(str_row)
        return rows

    def _apply_pisa_skip(self, rows: list[list[str]]) -> list[list[str]]:
        """Skip the title row for PISA 410 exports."""
        if _is_pisa_410(rows):
            logger.debug("PISA 410 format detected — skipping title row")
            return rows[1:]  # row[0] = title, row[1] onwards = header + data
        return rows

    # ------------------------------------------------------------------
    # CSV
    # ------------------------------------------------------------------

    def _parse_csv(self, path: Path) -> ExtractedDocument:
        content: str | None = None
        for encoding in ("utf-8-sig", "latin-1"):
            try:
                with open(path, encoding=encoding, newline="") as f:
                    content = f.read()
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError(f"Cannot decode {path} with supported encodings")
        rows = self._read_csv_rows(content)
        data_rows = self._apply_pisa_skip(rows)

        text_content = _rows_to_text(data_rows)
        metadata: dict[str, str] = {"row_count": str(len(data_rows))}

        return ExtractedDocument(
            source_path=str(path),
            format=DocumentFormat.CSV,
            text_content=text_content,
            structured_content={"tables": [data_rows]} if data_rows else None,
            metadata=metadata,
            embedded_image_count=0,
        )

    def _read_csv_rows(self, text: str) -> list[list[str]]:
        """Parse CSV text, auto-detecting semicolon vs comma delimiter."""
        delimiter = self._detect_delimiter(text)
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows: list[list[str]] = []
        for row in reader:
            str_row = [cell.strip() for cell in row]
            if any(cell for cell in str_row):
                rows.append(str_row)
        return rows

    @staticmethod
    def _detect_delimiter(text: str) -> str:
        """Return ';' if the first non-empty line has more semicolons than commas."""
        for line in text.splitlines():
            if line.strip():
                if line.count(";") >= line.count(","):
                    return ";"
                return ","
        return ","
